import  re
import  csv
import  xlrd
import  xlwt
import  os
from openpyxl import Workbook
from openpyxl import load_workbook
from xlutils.copy import copy

class bptest:

   def __init__(self):
        pass
               
   def check_if_contain(self, container, item):
        """return 0  if ``container`` does not contain ``item`` one or more times.
        return 1 if contains ``item`` one or more times.

        Works with strings, lists, and anything that supports Python's ``in``
        operator.

        See `Should Be Equal` for an explanation on how to override the default
        error message with arguments ``msg`` and ``values``.

        If ``ignore_case`` is given a true value (see `Boolean arguments`) and
        compared items are strings, it indicates that comparison should be
        case-insensitive. If the ``container`` is a list-like object, string
        items in it are compared case-insensitively. New option in Robot
        Framework 3.0.1.

        Examples:
        | Should Contain | ${output}    | PASS  |
        | Should Contain | ${some list} | value | msg=Failure! | values=False |
        | Should Contain | ${some list} | value | case_insensitive=True |
        """
        temp = container.find(item) 
        if temp == -1:
            return 0
        else:
            return 1
   
   
   def read_csv_file(self,filename):
        csv_dict=csv.DictReader(open(filename))
        return csv_dict
     
   def read_tsr_csvfile(self,location,filename):
        cf=location+'/'+filename
        csv_dict=self.read_csv_file(cf)
        return csv_dict
	 
   def xls_to_csv(self,filename,wsheet):
        wb=xlrd.open_workbook(filename)
        sh=wb.sheet_by_name(wsheet)
        curr_dir=os.getcwd()
        csvfile=curr_dir+'\\tmp.csv'
        a=open(csvfile,'wb')
        wr=csv.writer(a,quoting=csv.QUOTE_ALL)
        for rownum in xrange(sh.nrows):
             wr.writerow(sh.row_values(rownum))
        a.close()
                
   def read_xls_file(self,filename,worksheet):
         self.xls_to_csv(filename,worksheet)
         curr_dir=os.getcwd()
         cf=curr_dir+'\\tmp.csv'
         xls_data=self.read_csv_file(cf)
         return xls_data
   
   def get_match_filename(self,path,pattern):
   	    path1="//10.190.202.40/FIRMWARE/NG/"+path
   	    list = os.listdir(path1)
   	    list.sort()
   	    print list
   	    filename1=""
   	    for item in list:
   	    	if pattern in item:
   	    		filename1 = item
   	    		break 
   	    path2 =path1+"/"+filename1
   	    list1 = os.listdir(path2)
   	    print list1
   	    filename2=""
   	    for item in list1:
   	    	if ".sig" in item and "d" not in item:
   	    		filename2 = item
   	    		break
   	    download_path=path+"/"+filename1+"/"+filename2
   	    print download_path
   	    return download_path
	
   def create_IMIX_reportFile(self,path,filename):
   	    csvfile = path+'/'+filename
   	    a = open(csvfile,'wb')
   	    fileheader=["Throughput(Mbps)","FrameRate(kfps)"]
   	    dict_writer = csv.DictWriter(a,fileheader)
   	    dict_writer.writeheader()
   	    a.close()	
		
   def get_IMIXResult_2ports(self,path,filename):
   	    result=[0,0]
   	    csvfile = path+'/'+filename
   	    data=xlrd.open_workbook(csvfile,'r')
   	    table=data.sheets()[0]
   	    nrows=table.nrows
   	    ncols=table.ncols
#search all row, get Frame date rate value
   	    for curr_row in range(nrows):
   	    	line = table.row_values(curr_row)
   	    	line1=str(line[0])
   	    	str1='7.10.4. Frame Data Rate'
   	    	str2='7.10.4.1.'
   	    	if line1.find(str1) != -1: 
   	    		line_start = curr_row
   	    		
   	    	if line1.find(str2)!= -1:
   	    		line_end = curr_row
   	    		break
   	    print('line_start line_end is %d %d'%(line_start,line_end)) 
   	    value = 0
   	    row_start = line_start +7
   	    j = int((line_end-row_start)/2)
   	    i = row_start + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell_value(index,2).replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge = float(value/50)
   	    print('averge Frame Data Rate value is %d'%averge) 
   	    result[0]=averge

#serch for  Frame Rate
   	    for curr_row in range(nrows):
   	    	line = table.row_values(curr_row)
   	    	line1=str(line[0])
   	    	str4='7.10.2. Frame Rate'
   	    	str5='7.10.2.1.'
   	    	if line1.find(str4) != -1: 
   	    		line_start_FrameRate = curr_row
   	    			
   	    	if line1.find(str5)!= -1:
   	    		line_end_FrameRate = curr_row
   	    		break
   	    print('line_start line_end is %d %d'%(line_start_FrameRate,line_end_FrameRate)) 
   	    value = 0
   	    row_start_FrameRate = line_start_FrameRate +4
   	    j = int((line_end_FrameRate-row_start_FrameRate)/2)
   	    i = row_start_FrameRate + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell_value(index,2).replace(',','')
   	    	tmp=float(tmp)
   	    	value = value + tmp
   	    averge1 = float(value/50)
   	    print('averge Frame  Rate value is %d'%averge1) 
   	    result[1]=float(averge1/1000)
   	    return result
   
   def create_RFC2544_reportFile(self,path,filename):
   	    csvfile = path+'/'+filename
   	    a = open(csvfile,'wb')
   	    fileheader=["FrameSize(Byte)","Throughput(Mbps)","MinLatency(us)","FrameRate(kfps)"]
   	    dict_writer = csv.DictWriter(a,fileheader)
   	    dict_writer.writeheader()
   	    a.close()
	
   def create_TSR_File(self,path,filename):
   	    csvfile = path+'/'+filename
   	    a = open(csvfile,'wb')
   	    fileheader=["Address Objects","Address Groups","Address Group Depth","AO Per Address Group","Service Objects","Service Groups","Service Group Depth","SO Per Service Group","NAT Policies","VLANs","Zones","Local Users","User Groups","User Logins","Guest Users","Rules Per Table","DHCP Ranges","SonicPoint Per Interface","S2S VPN Policies","GroupVPN Policies","Static Routes","Tunnel interface","Static ARP Entries","Schedule Objects","Application Firewall Policies","Application Firewall APP Objects","Application Firewall Actions","Application Firewall Email User Objects","VPN Phase 1 SA's","VPN Phase 2 SA's","SSLVPN Connections","Botnet Filters","GeoIP Filters","Static Routes for FQDN","REST API Agents","REST API User Logins","DPI-SSH","Dynamic Address Objects",]
   	    dict_writer = csv.DictWriter(a,fileheader)
   	    dict_writer.writeheader()
   	    a.close()
   	 
   def get_RFC2544Result_2ports(self,path,filename,lenth):
   	    result=[lenth,0,0,0]
   	    csvfile = path+'/'+filename
   	    data=xlrd.open_workbook(csvfile,'r')
   	    table=data.sheets()[0]
   	    nrows=table.nrows
   	    ncols=table.ncols
#search all row, get Frame date rate value
   	    for curr_row in range(nrows):
   	    	line = table.row_values(curr_row)
   	    	line1=str(line[0])
   	    	str1='7.10.4. Frame Data Rate'
   	    	str2='7.10.4.1.'
   	    	if line1.find(str1) != -1: 
   	    		line_start = curr_row
   	    		
   	    	if line1.find(str2)!= -1:
   	    		line_end = curr_row
   	    		break
   	    print('line_start line_end is %d %d'%(line_start,line_end)) 
   	    value = 0
   	    row_start = line_start +7
   	    j = int((line_end-row_start)/2)
   	    i = row_start + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell_value(index,2).replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge = float(value/50)
   	    print('averge Frame Data Rate value is %d'%averge) 
   	    result[1]=averge
#search for Minimum Frame Latency
   	    for curr_row in range(nrows):
   	    	line = table.row_values(curr_row)
   	    	line1=str(line[0])
   	    	str3='7.6. Frame Latency Summary'
   	    	if line1.find(str3) != -1: 
   	    		line_Latency = curr_row
   	    		break
   	    line_Latency=line_Latency+3
   	    result[2]=table.cell_value(line_Latency,1)
#serch for  Frame Rate
   	    for curr_row in range(nrows):
   	    	line = table.row_values(curr_row)
   	    	line1=str(line[0])
   	    	str4='7.10.2. Frame Rate'
   	    	str5='7.10.2.1.'
   	    	if line1.find(str4) != -1: 
   	    		line_start_FrameRate = curr_row
   	    			
   	    	if line1.find(str5)!= -1:
   	    		line_end_FrameRate = curr_row
   	    		break
   	    print('line_start line_end is %d %d'%(line_start_FrameRate,line_end_FrameRate)) 
   	    value = 0
   	    row_start_FrameRate = line_start_FrameRate +4
   	    j = int((line_end_FrameRate-row_start_FrameRate)/2)
   	    i = row_start_FrameRate + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell_value(index,2).replace(',','')
   	    	tmp=float(tmp)
   	    	value = value + tmp
   	    averge1 = float(value/50)
   	    print('averge Frame  Rate value is %d'%averge1) 
   	    result[3]=float(averge1/1000)
   	    return result
   
   def get_IMIXResult_4ports(self,path,filename):
   	    result=[0,0]
   	    xlsxfile = path+'/'+filename
   	    data=load_workbook(xlsxfile,'r')
   	    sheetnames = data.get_sheet_names()
   	    table=data.get_sheet_by_name(sheetnames[0])
   	    nrows=table.max_row
   	    ncols=table.max_column
#search all row, get Frame date rate value
   	    x=0
   	    y=0
   	    for row in table.rows:
   	      line = row[0].value
   	      line1=str(line)
   	      str1='7.6.4. Frame Data Rate'
   	      str2='7.6.4.1.'
   	      x=x+1
   	      y=y+1
   	      if line1.find(str1) != -1: 
   	       line_start = x
   	      
   	      if line1.find(str2) != -1:
   	       line_end = y
   	       break
   	    print('line_start line_end is %d %d'%(line_start,line_end))
   	    value = 0
   	    row_start = line_start +7
   	    j = int((line_end-row_start)/2)
   	    i = row_start + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell(index,3).value
   	    	tmp = tmp.replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge = float(value/50)
   	    print('averge Frame Data Rate value is %f'%averge) 
   	    result[0]=averge

#serch for  Frame Rate
   	    x=0
   	    y=0
   	    for row in table.rows:
   	    	line = row[0].value
   	    	line1=str(line)
   	    	str4='7.6.2. Frame Rate'
   	    	str5='7.6.2.1. Frame Rate'
   	    	x=x+1
   	    	y=y+1
   	    	if line1.find(str4) != -1: 
   	    		line_start_FrameRate = x
   	    			
   	    	if line1.find(str5)!= -1:
   	    		line_end_FrameRate = y
   	    		break
   	    print('line_start line_end is %d %d'%(line_start_FrameRate,line_end_FrameRate)) 
   	    value = 0
   	    row_start_FrameRate = line_start_FrameRate +4
   	    j = int((line_end_FrameRate-row_start_FrameRate)/2)
   	    i = row_start_FrameRate + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell(index,3).value
   	    	tmp = tmp.replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge1 = float(value/50)
   	    print('averge Frame  Rate value is %d'%averge1) 
   	    result[1]=float(averge1/1000)
   	    return result
   
   def csv_to_xls(self,path,filename):
   	    csvfile = path+'/'+filename+'.csv'
   	    with open(csvfile,'r') as f:
   	    	read = csv.reader(f)
   	    	workbook = xlwt.Workbook()
   	    	sheet = workbook.add_sheet('data')  
   	    	l = 0
   	    	for line in read:
   	    		r = 0
   	    		for i in line:
   	    			sheet.write(l, r, i)  
   	    			r = r + 1
   	    		l = l + 1
   	    	csvfile1 = path+'/'+filename+'.xls'
   	    	workbook.save(csvfile1)	  	
   
   def write_to_csvFile(self,path,filename,*arg):
   	    csvfile = path+'/'+filename+'.csv'
   	    a = open(csvfile,'ab')
   	    dict_writer = csv.writer(a)
   	    dict_writer.writerow((arg))
   	    a.close()	  

   def csv_to_xlsx(self,path,filename):
   	    csvfile = path+'/'+filename+'.csv'
   	    with open(csvfile,'r') as f:
   	    	read = csv.reader(f)
   	    	workbook = Workbook()
   	    	sheet = workbook.active
   	    	sheet.title ='data'
   	    	l = 1
   	    	for line in read:
   	    		r = 1
   	    		for i in line:
   	    			sheet.cell(row=l,column=r).value=i
   	    			r = r + 1
   	    		l = l + 1
   	    	csvfile1 = path+'/'+filename+'.xlsx'
   	    	workbook.save(csvfile1)  
   	    	
   def get_RFC2544Result_4ports(self,path,filename,lenth):
   	    result=[lenth,0,0,0]
   	    xlsxfile = path+'/'+filename
   	    data=load_workbook(xlsxfile,'r')
   	    sheetnames = data.get_sheet_names()
   	    table=data.get_sheet_by_name(sheetnames[0])
   	    nrows=table.max_row
   	    ncols=table.max_column
#search all row, get Frame date rate value
   	    x=0
   	    y=0
   	    for row in table.rows:
   	      line = row[0].value
   	      line1=str(line)
   	      str1='7.6.4. Frame Data Rate'
   	      str2='7.6.4.1.'
   	      x=x+1
   	      y=y+1
   	      if line1.find(str1) != -1: 
   	       line_start = x
   	      
   	      if line1.find(str2) != -1:
   	       line_end = y
   	       break
   	    print('line_start line_end is %d %d'%(line_start,line_end))
   	    value = 0
   	    row_start = line_start +7
   	    j = int((line_end-row_start)/2)
   	    i = row_start + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell(index,3).value
   	    	tmp = tmp.replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge = float(value/50)
   	    print('averge Frame Data Rate value is %f'%averge) 
   	    result[1]=averge
#search for Minimum Frame Latency
   	    x=0
   	    for row in table.rows:
   	    	line = row[0].value
   	    	line1=str(line)
   	    	str3='7.2. Frame Latency Summary'
   	    	x=x+1
   	    	if line1.find(str3) != -1: 
   	    		line_Latency = x
   	    		break
   	    line_Latency=line_Latency+3
   	    result[2]=table.cell(line_Latency,2).value
#serch for  Frame Rate
   	    x=0
   	    y=0
   	    for row in table.rows:
   	    	line = row[0].value
   	    	line1=str(line)
   	    	str4='7.6.2. Frame Rate'
   	    	str5='7.6.2.1. Frame Rate'
   	    	x=x+1
   	    	y=y+1
   	    	if line1.find(str4) != -1: 
   	    		line_start_FrameRate = x
   	    			
   	    	if line1.find(str5)!= -1:
   	    		line_end_FrameRate = y
   	    		break
   	    print('line_start line_end is %d %d'%(line_start_FrameRate,line_end_FrameRate)) 
   	    value = 0
   	    row_start_FrameRate = line_start_FrameRate +4
   	    j = int((line_end_FrameRate-row_start_FrameRate)/2)
   	    i = row_start_FrameRate + j
   	    for index in range(i, i+50):
   	    	tmp = table.cell(index,3).value
   	    	tmp = tmp.replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge1 = float(value/50)
   	    print('averge Frame  Rate value is %d'%averge1) 
   	    result[3]=float(averge1/1000)
   	    return result
   	    
   def create_UTM_reportFile(self,path,filename):
   	    workbook = xlwt.Workbook()
   	    sheet = workbook.add_sheet('UTM Result')
   	    sheet.write(0,0,"UTM OFF(Megabits/s)")
   	    sheet.write(0,1,"IPS - Max Security(Megabits/s)")
   	    sheet.write(0,2,"IPS - Perf Optimized(Megabits/s)")
   	    sheet.write(0,3,"GAV -Max Security(Megabits/s)")
   	    sheet.write(0,4,"GAV - Perf Optimized(Megabits/s)")
   	    sheet.write(0,5,"Anti-Spy -Max Sec(Megabits/s)")
   	    sheet.write(0,6,"Anti-Spy - Perf Opt(Megabits/s)")
   	    sheet.write(0,7,"All -Max Security(Megabits/s)")
   	    sheet.write(0,8,"All - Perf Optimized(Megabits/s)")
   	    csvfile1 = path+'/'+filename
   	    workbook.save(csvfile1)
   	    
   def write_to_xlsFile(self,path,filename,row,col,value):
   	    file = path+'/'+filename
   	    data=xlrd.open_workbook(file,'rb')
   	    wb=copy(data)
   	    ws=wb.get_sheet(0)
   	    row=int(row)
   	    col=int(col)
   	    ws.write(row,col,value)
   	    wb.save(file)
   	    wb.save(file)
   	    
   def get_UTMResult_2ports(self,path,filename):
   	    result=0
   	    csvfile = path+'/'+filename
   	    data=xlrd.open_workbook(csvfile,'r')
   	    table=data.sheets()[0]
   	    nrows=table.nrows
   	    ncols=table.ncols
#search all row, get Frame date rate value
   	    for curr_row in range(nrows):
   	    	line = table.row_values(curr_row)
   	    	line1=str(line[0])
   	    	str1=re.compile('7.23.24. Frame Data Rate')
   	    	str2=re.compile('7.23.24.1. Frame Data Rate')
   	    	if str1.findall(line1): 
   	    		line_start = curr_row
   	    		print str1.findall(line1)
   	    		
   	    	if str2.findall(line1):
   	    		line_end = curr_row
   	    		print str2.findall(line1)
   	    		break
   	    print('line_start line_end is %d %d'%(line_start,line_end)) 
   	    value = 0
   	    row_start = line_start +7
   	    j = int((line_end-row_start)/2)
   	    i = row_start + j
   	    for index in range(i, i+30):
   	    	tmp = table.cell_value(index,2).replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge = float(value/30)
   	    print('averge Frame Data Rate value is %d'%averge) 
   	    result=averge
   	    return result
   	    
   def get_UTMResult_4ports(self,path,filename):
   	    result=0
   	    csvfile = path+'/'+filename
   	    data=xlrd.open_workbook(csvfile,'r')
   	    table=data.sheets()[0]
   	    nrows=table.nrows
   	    ncols=table.ncols
#search all row, get Frame date rate value
   	    for curr_row in range(nrows):
   	    	line = table.row_values(curr_row)
   	    	line1=str(line[0])
   	    	str1=re.compile('7.18.24. Frame Data Rate')
   	    	str2=re.compile('7.18.24.1. Frame Data Rate')
   	    	if str1.findall(line1): 
   	    		line_start = curr_row
   	    		print str1.findall(line1)
   	    		
   	    	if str2.findall(line1):
   	    		line_end = curr_row
   	    		print str2.findall(line1)
   	    		break
   	    print('line_start line_end is %d %d'%(line_start,line_end)) 
   	    value = 0
   	    row_start = line_start +7
   	    j = int((line_end-row_start)/2)
   	    i = row_start + j
   	    for index in range(i, i+30):
   	    	tmp = table.cell_value(index,2).replace(',','')
   	    	tmp1=float(tmp)
   	    	value = value + tmp1
   	    averge = float(value/30)
   	    print('averge Frame Data Rate value is %d'%averge) 
   	    result=averge
   	    return result
   	    
   def create_folder(self,targetPath):
   	    if not os.path.exists(targetPath):
   	    	os.makedirs(targetPath)
   	    else:
   	    	print('The folder has already existed!')